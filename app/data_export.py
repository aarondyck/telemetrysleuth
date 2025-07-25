#!/usr/bin/env python3
"""
Data Export Module for Telemetry Sleuth.

This module provides functionality to export call records in various formats
including CSV, JSON, and Excel.
"""

import csv
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from io import StringIO, BytesIO
import zipfile
from .database import DatabaseManager

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel export will be disabled.")


class DataExporter:
    """Handles data export in various formats."""
    
    def __init__(self, db_manager: Optional[DatabaseManager] = None):
        self.db_manager = db_manager or DatabaseManager()
        
        # Define field mappings for export
        self.field_mappings = {
            'id': 'Record ID',
            'call_start_time': 'Call Start Time',
            'connected_time': 'Duration (seconds)',
            'ring_time': 'Ring Time (seconds)',
            'caller': 'Caller Number',
            'direction': 'Direction',
            'called_number': 'Called Number',
            'dialed_number': 'Dialed Number',
            'account_code': 'Account Code',
            'is_internal': 'Internal Call',
            'call_id': 'Call ID',
            'continuation': 'Continuation',
            'party1_device': 'Party 1 Device',
            'party1_name': 'Party 1 Name',
            'party2_device': 'Party 2 Device',
            'party2_name': 'Party 2 Name',
            'hold_time': 'Hold Time (seconds)',
            'park_time': 'Park Time (seconds)',
            'authorization_valid': 'Authorization Valid',
            'authorization_code': 'Authorization Code',
            'call_charge': 'Call Charge',
            'call_units': 'Call Units',
            'cost_per_unit': 'Cost Per Unit',
            'mark_up': 'Mark Up',
            'currency': 'Currency',
            'smdr_record_time': 'SMDR Record Time'
        }
    
    def export_to_csv(self, records: List[Dict[str, Any]], include_headers: bool = True) -> str:
        """Export records to CSV format."""
        if not records:
            return ""
        
        output = StringIO()
        
        # Get field names from the first record
        fieldnames = list(records[0].keys())
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        
        if include_headers:
            # Write headers with friendly names
            header_row = {field: self.field_mappings.get(field, field.title()) for field in fieldnames}
            writer.writerow(header_row)
        
        # Write data rows
        for record in records:
            # Convert datetime objects to strings
            processed_record = {}
            for key, value in record.items():
                if isinstance(value, datetime):
                    processed_record[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    processed_record[key] = ''
                elif isinstance(value, bool):
                    processed_record[key] = 'Yes' if value else 'No'
                else:
                    processed_record[key] = str(value)
            
            writer.writerow(processed_record)
        
        return output.getvalue()
    
    def export_to_json(self, records: List[Dict[str, Any]], pretty: bool = True) -> str:
        """Export records to JSON format."""
        if not records:
            return "[]"
        
        # Process records for JSON serialization
        processed_records = []
        for record in records:
            processed_record = {}
            for key, value in record.items():
                if isinstance(value, datetime):
                    processed_record[key] = value.isoformat()
                else:
                    processed_record[key] = value
            processed_records.append(processed_record)
        
        if pretty:
            return json.dumps(processed_records, indent=2, ensure_ascii=False)
        else:
            return json.dumps(processed_records, ensure_ascii=False)
    
    def export_to_excel(self, records: List[Dict[str, Any]], sheet_name: str = "Call Records") -> bytes:
        """Export records to Excel format."""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        if not records:
            # Create empty workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["No data available"])
            
            output = BytesIO()
            wb.save(output)
            return output.getvalue()
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Get field names
        fieldnames = list(records[0].keys())
        
        # Create header row with friendly names
        headers = [self.field_mappings.get(field, field.title()) for field in fieldnames]
        ws.append(headers)
        
        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row_num, record in enumerate(records, 2):
            for col_num, field in enumerate(fieldnames, 1):
                value = record[field]
                
                # Format values for Excel
                if isinstance(value, datetime):
                    ws.cell(row=row_num, column=col_num, value=value)
                elif value is None:
                    ws.cell(row=row_num, column=col_num, value="")
                elif isinstance(value, bool):
                    ws.cell(row=row_num, column=col_num, value="Yes" if value else "No")
                else:
                    ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for col_num, field in enumerate(fieldnames, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(self.field_mappings.get(field, field.title()))
            
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # Set column width (with some padding)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_records_by_filters(self, 
                                 format_type: str,
                                 date_from: Optional[datetime] = None,
                                 date_to: Optional[datetime] = None,
                                 direction: Optional[str] = None,
                                 is_internal: Optional[bool] = None,
                                 caller: Optional[str] = None,
                                 called: Optional[str] = None,
                                 limit: Optional[int] = None) -> tuple:
        """
        Export records based on filters.
        
        Returns:
            tuple: (data, content_type, filename)
        """
        try:
            # Get records from database
            records = self.db_manager.search_call_records(
                date_from=date_from,
                date_to=date_to,
                direction=direction,
                is_internal=is_internal,
                caller=caller,
                called=called,
                limit=limit or 10000  # Default limit for exports
            )
            
            if not records:
                logger.warning("No records found for export")
                return None, None, None
            
            # Convert records to dictionaries
            record_dicts = []
            for record in records:
                record_dict = {}
                for column in record.__table__.columns:
                    record_dict[column.name] = getattr(record, column.name)
                record_dicts.append(record_dict)
            
            # Generate timestamp for filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Export based on format
            if format_type.lower() == 'csv':
                data = self.export_to_csv(record_dicts)
                content_type = 'text/csv'
                filename = f'call_records_{timestamp}.csv'
                
            elif format_type.lower() == 'json':
                data = self.export_to_json(record_dicts)
                content_type = 'application/json'
                filename = f'call_records_{timestamp}.json'
                
            elif format_type.lower() == 'excel':
                if not EXCEL_AVAILABLE:
                    raise ValueError("Excel export not available. Install openpyxl.")
                
                data = self.export_to_excel(record_dicts)
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                filename = f'call_records_{timestamp}.xlsx'
                
            else:
                raise ValueError(f"Unsupported export format: {format_type}")
            
            logger.info(f"Exported {len(record_dicts)} records to {format_type.upper()} format")
            return data, content_type, filename
            
        except Exception as e:
            logger.error(f"Error exporting records: {e}")
            raise
    
    def create_export_archive(self, 
                             date_from: Optional[datetime] = None,
                             date_to: Optional[datetime] = None,
                             formats: List[str] = None) -> tuple:
        """
        Create a ZIP archive containing exports in multiple formats.
        
        Returns:
            tuple: (zip_data, content_type, filename)
        """
        if formats is None:
            formats = ['csv', 'json']
            if EXCEL_AVAILABLE:
                formats.append('excel')
        
        try:
            # Create ZIP file in memory
            zip_buffer = BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Add exports in each format
                for format_type in formats:
                    try:
                        data, _, filename = self.export_records_by_filters(
                            format_type=format_type,
                            date_from=date_from,
                            date_to=date_to
                        )
                        
                        if data:
                            if isinstance(data, str):
                                zip_file.writestr(filename, data.encode('utf-8'))
                            else:
                                zip_file.writestr(filename, data)
                    
                    except Exception as e:
                        logger.error(f"Error adding {format_type} to archive: {e}")
                        # Add error file instead
                        error_msg = f"Error generating {format_type} export: {str(e)}"
                        zip_file.writestr(f"error_{format_type}.txt", error_msg)
                
                # Add metadata file
                metadata = {
                    'export_timestamp': datetime.now().isoformat(),
                    'date_from': date_from.isoformat() if date_from else None,
                    'date_to': date_to.isoformat() if date_to else None,
                    'formats_included': formats,
                    'exported_by': 'Telemetry Sleuth'
                }
                
                zip_file.writestr('export_metadata.json', 
                                json.dumps(metadata, indent=2))
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'call_records_export_{timestamp}.zip'
            
            return zip_buffer.getvalue(), 'application/zip', filename
            
        except Exception as e:
            logger.error(f"Error creating export archive: {e}")
            raise
    
    def get_export_summary(self, 
                          date_from: Optional[datetime] = None,
                          date_to: Optional[datetime] = None) -> Dict[str, Any]:
        """Get summary information for potential export."""
        try:
            # Get record count
            records = self.db_manager.search_call_records(
                date_from=date_from,
                date_to=date_to,
                limit=1  # Just need count
            )
            
            total_count = self.db_manager.get_record_count(
                date_from=date_from,
                date_to=date_to
            )
            
            # Get date range of available data
            date_range = self.db_manager.get_date_range()
            
            return {
                'total_records': total_count,
                'date_range': {
                    'earliest': date_range['earliest'].isoformat() if date_range['earliest'] else None,
                    'latest': date_range['latest'].isoformat() if date_range['latest'] else None
                },
                'available_formats': ['csv', 'json'] + (['excel'] if EXCEL_AVAILABLE else []),
                'estimated_file_sizes': {
                    'csv': f"{total_count * 200} bytes (estimated)",  # Rough estimate
                    'json': f"{total_count * 400} bytes (estimated)",
                    'excel': f"{total_count * 300} bytes (estimated)" if EXCEL_AVAILABLE else "Not available"
                }
            }
            
        except Exception as e:
            logger.error(f"Error getting export summary: {e}")
            return {
                'error': str(e),
                'total_records': 0,
                'available_formats': []
            }